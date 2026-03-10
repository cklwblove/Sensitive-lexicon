#!/usr/bin/env python3
# 汇总 ThirdPartyCompatibleFormats/TrChat 与 Vocabulary 词库，导出为 Excel

import json
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("请先安装: pip install openpyxl")
    raise

ROOT = Path(__file__).resolve().parent
TRCHAT_JSON = ROOT / "ThirdPartyCompatibleFormats" / "TrChat" / "SensitiveLexicon.json"
VOCAB_DIR = ROOT / "Vocabulary"
OUT_EXCEL = ROOT / "敏感词库汇总.xlsx"


def load_trchat():
    with open(TRCHAT_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("words") or []


def load_vocabulary():
    word_sources = defaultdict(set)
    for p in sorted(VOCAB_DIR.glob("*.txt")):
        cat = p.stem
        with open(p, "r", encoding="utf-8") as f:
            for line in f:
                w = line.strip()
                if w:
                    word_sources[w].add(cat)
    return word_sources


def main():
    # 词 -> 来源集合
    word_sources = defaultdict(set)
    trchat_words = load_trchat()
    for w in trchat_words:
        if w:
            word_sources[w].add("TrChat")
    for w, sources in load_vocabulary().items():
        word_sources[w].update(sources)
    # 排序：先按来源数量（多来源靠前），再按敏感词
    rows = []
    for w, sources in word_sources.items():
        src_str = "、".join(sorted(sources))
        rows.append((w, src_str))
    rows.sort(key=lambda x: (-len(x[1]), x[0]))
    # 写 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "敏感词库"
    ws.append(["序号", "敏感词", "来源"])
    for i, (w, src) in enumerate(rows, 1):
        ws.append([i, w, src])
    # 表头样式
    for c in "ABC":
        ws[f"{c}1"].font = Font(bold=True)
        ws[f"{c}1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 36
    wb.save(OUT_EXCEL)
    print(f"已导出: {OUT_EXCEL}，共 {len(rows)} 条敏感词")


if __name__ == "__main__":
    main()
